# -*-coding:utf-8-*-
from openpyxl import load_workbook, Workbook
import Tkinter
from Tkinter import *
from utils import *
from sheetHandler import *
from cmdInterface import *


if __name__ == "__main__":
    wb = load_workbook("test.xlsx")
    # show_sheet_names(wb.sheetnames)
    # select_mode()
    sheet = wb["totalRune"]
    print(sheet.max_row)
    s = 'A' + str(sheet.max_row)
    print(sheet[s].value)

    # wb = Workbook()
    # # grab the active worksheet
    # ws = wb.active
    # # Data can be assigned directly to cells
    # ws['A1'] = 42
    # # Rows can also be appended
    # ws.append([1, 2, 3])
    # # Python types will automatically be converted
    # import datetime
    # ws['A2'] = datetime.datetime.now()
    # # Save the file
    # wb.save("sample.xlsx")
